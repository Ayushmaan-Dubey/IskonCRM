import base64
import secrets
from datetime import datetime
from functools import wraps

from flask import (Blueprint, render_template, redirect, url_for, flash,
                   request, abort)
from flask_login import current_user
from werkzeug.security import generate_password_hash

from .models import (User, Person, GuestIntake, Sponsorship,
                     InventoryItem, InventoryAuditLog, MessageTemplate,
                     FactSheet, PaymentSettings, SecurityAuditLog)
from . import db
from .audit import log_event

try:
    from .email_utils import send_new_user_email
except Exception:
    def send_new_user_email(to_email, temp_password):
        return False, 'email_utils not available'

admin = Blueprint('admin', __name__)

ROLE_OPTIONS = [
    ('member', 'Member'),
    ('user', 'User'),
    ('sales', 'Sales'),
    ('sunday_school_teacher', 'Sunday School Teacher'),
    ('congregational_development', 'Congregational Development'),
    ('admins', 'Admin'),
    ('super_admin', 'Super Admin'),
]

HPO_LOCATIONS = [
    'Round Rock', 'Austin', 'Cedar Park', 'Georgetown',
    'Liberty Hill', 'Pflugerville', 'Leander', 'Other',
]


def _parse_date(value):
    if not value:
        return None
    try:
        return datetime.strptime(value, '%Y-%m-%d').date()
    except ValueError:
        return None


# ─── Decorators ──────────────────────────────────────────────

def admin_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if not current_user.is_authenticated:
            flash('Please log in as admin to access this page.', 'info')
            return redirect(url_for('auth.login'))
        if not getattr(current_user, 'is_admin', False):
            flash('Access denied.', 'danger')
            return redirect(url_for('views.home'))
        return f(*args, **kwargs)
    return decorated


def super_admin_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if not current_user.is_authenticated:
            flash('Please log in as admin to access this page.', 'info')
            return redirect(url_for('auth.login'))
        if not getattr(current_user, 'is_super_admin', False):
            flash('This page requires Super Admin access.', 'danger')
            return redirect(url_for('admin.admin_index'))
        return f(*args, **kwargs)
    return decorated


# ─── User Management ─────────────────────────────────────────

@admin.route('/admin')
@admin_required
def admin_index():
    users = User.query.all() if hasattr(User, 'query') else []
    return render_template('admin/users.html', users=users, role_options=ROLE_OPTIONS)


@admin.route('/admin/users')
@admin_required
def admin_users():
    users = User.query.order_by(User.id.desc()).all() if hasattr(User, 'query') else []
    return render_template('admin/users.html', users=users, role_options=ROLE_OPTIONS)


@admin.route('/admin/users/<int:user_id>/role', methods=['POST'])
@admin_required
def admin_update_user_role(user_id):
    user = User.query.get_or_404(user_id)
    selected_role = request.form.get('role', 'user')
    valid_roles = {value for value, _label in ROLE_OPTIONS}
    if selected_role not in valid_roles:
        flash('Invalid role selected.', 'warning')
        return redirect(url_for('admin.admin_users'))
    # Only super_admin can assign super_admin role
    if selected_role == 'super_admin' and not current_user.is_super_admin:
        flash('Only a Super Admin can assign the Super Admin role.', 'danger')
        return redirect(url_for('admin.admin_users'))
    old_role = user.role
    user.role = selected_role
    user.is_admin = selected_role in ('admins', 'super_admin')
    db.session.commit()
    log_event('role_changed', f'{user.email}: {old_role} -> {selected_role}', actor=current_user,
              target_type='User', target_id=user.id)
    flash(f'Updated role for {user.email}.', 'success')
    return redirect(url_for('admin.admin_users'))


@admin.route('/admin/create-user', methods=['GET', 'POST'])
@admin_required
def admin_create_user():
    errors = []
    if request.method == 'POST':
        email = request.form.get('email', '').strip()
        legal_name = request.form.get('legalName', '').strip()
        username = request.form.get('username', '').strip()
        phone_number = request.form.get('phoneNumber', '').strip()
        firstName = legal_name.split()[0] if legal_name else request.form.get('firstName', '').strip()
        lastName = request.form.get('lastName', '').strip()
        first_time = request.form.get('first_time_at_temple', 'no') == 'yes'
        contact_date = _parse_date(request.form.get('contact_date', '').strip())
        area = request.form.get('area', '').strip() or None
        interests_list = request.form.getlist('interests')
        interests = ', '.join(interests_list) if interests_list else None
        event_source = request.form.get('event_source', '').strip() or None

        if db is not None:
            if firstName and lastName:
                name_conflict = User.query.filter_by(first_name=firstName, last_name=lastName).first()
                if name_conflict:
                    flash('User already exists. Please check existing users.', 'warning')
                    return redirect(url_for('admin.admin_users'))

        if not email or not firstName:
            errors.append('Please fill out all required fields.')
        if not legal_name or not username or not phone_number:
            errors.append('Legal name, username, and phone number are required.')

        if db is not None:
            if User.query.filter_by(email=email).first():
                errors.append('An account with that email already exists.')
            if User.query.filter_by(username=username).first():
                errors.append('That username is already taken.')

        if errors:
            return render_template('admin/create_user.html', errors=errors, form=request.form)

        if db is not None:
            temp_password = secrets.token_urlsafe(8)
            new_user = User(
                email=email,
                first_name=firstName,
                legal_name=legal_name,
                username=username,
                phone_number=phone_number,
                last_name=lastName,
                password=generate_password_hash(temp_password, method='pbkdf2:sha256'),
                is_admin=False,
                role='user',
                must_change_password=True,
                first_time_at_temple=first_time,
                contact_date=contact_date,
                area=area,
                interests=interests,
                event_source=event_source,
                created_by_admin=getattr(current_user, 'first_name', None),
            )
            db.session.add(new_user)
            db.session.commit()
            log_event('admin_action', f'Created user {email}.', actor=current_user,
                      target_type='User', target_id=new_user.id)
            sent, err = send_new_user_email(email, temp_password)
            if not sent:
                flash(f'User created but email failed: {err}', 'warning')
            else:
                flash('User created and email sent successfully.', 'success')

        return redirect(url_for('admin.admin_users'))

    return render_template('admin/create_user.html')


# ─── Unified People / Member Database (SRS §2) ───────────────

@admin.route('/admin/people')
@admin_required
def admin_people():
    tag_filter = request.args.get('tag', '')
    query = Person.query
    if tag_filter:
        query = query.filter(Person.status_tags.contains(tag_filter))
    people = query.order_by(Person.full_name).all()
    return render_template('admin/people.html', people=people, tag_filter=tag_filter)


@admin.route('/admin/people/<int:person_id>/tags', methods=['POST'])
@admin_required
def admin_people_update_tags(person_id):
    person = Person.query.get_or_404(person_id)
    new_tags = request.form.getlist('tags')
    person.status_tags = ','.join(new_tags)
    db.session.commit()
    flash(f'Updated tags for {person.full_name}.', 'success')
    return redirect(url_for('admin.admin_people'))


@admin.route('/admin/people/<int:person_id>/invite', methods=['POST'])
@admin_required
def admin_people_invite_donation(person_id):
    """Manually trigger a donation invitation email."""
    person = Person.query.get_or_404(person_id)
    if not person.email:
        flash('This person has no email address on record.', 'warning')
        return redirect(url_for('admin.admin_people'))
    try:
        from .email_utils import send_not_yet_paid_email
        settings = PaymentSettings.query.first()
        from flask import request as req
        form_url = req.host_url.rstrip('/') + url_for('views.donate')
        sent, err = send_not_yet_paid_email(person.email, person.full_name, settings, form_url)
        if sent:
            flash(f'Donation invitation sent to {person.email}.', 'success')
        else:
            flash(f'Failed to send email: {err}', 'warning')
    except Exception as e:
        flash(f'Email error: {e}', 'danger')
    return redirect(url_for('admin.admin_people'))


# ─── Spreadsheet Import (Super Admin only) ───────────────────

IMPORT_COLUMNS = ['Full Name', 'Email', 'Phone', 'Tags', 'Notes']


def _find_or_create_person_for_import(full_name, email, phone_number):
    """Returns (person, was_created)."""
    person = None
    if email:
        person = Person.query.filter_by(email=email).first()
    if person is None and phone_number:
        person = Person.query.filter(
            Person.phone_number == phone_number,
            Person.email.is_(None)
        ).first()
    if person is not None:
        return person, False
    person = Person(full_name=full_name, email=email or None, phone_number=phone_number or None)
    db.session.add(person)
    db.session.flush()
    return person, True


@admin.route('/admin/import', methods=['GET', 'POST'])
@super_admin_required
def admin_import_people():
    if request.method == 'POST':
        try:
            from openpyxl import load_workbook
        except ImportError:
            flash('openpyxl is not installed. Run: pip install openpyxl', 'danger')
            return render_template('admin/import_people.html', columns=IMPORT_COLUMNS)

        file = request.files.get('spreadsheet')
        if not file or not file.filename:
            flash('Please select a .xlsx file to upload.', 'warning')
            return render_template('admin/import_people.html', columns=IMPORT_COLUMNS)

        wb = load_workbook(file, read_only=True, data_only=True)
        ws = wb.active

        rows = ws.iter_rows(values_only=True)
        header = [str(c).strip().lower() if c else '' for c in next(rows, [])]

        def col(name):
            try:
                return header.index(name)
            except ValueError:
                return None

        idx_name = col('full name')
        if idx_name is None:
            flash('The spreadsheet must have a "Full Name" column. Download the template for the expected format.', 'danger')
            return render_template('admin/import_people.html', columns=IMPORT_COLUMNS)

        idx_email = col('email')
        idx_phone = col('phone')
        idx_tags = col('tags')
        idx_notes = col('notes')

        def cell(row, idx):
            if idx is None or idx >= len(row) or row[idx] is None:
                return ''
            return str(row[idx]).strip()

        created = updated = skipped = 0
        for row in rows:
            full_name = cell(row, idx_name)
            if not full_name:
                skipped += 1
                continue
            email = cell(row, idx_email) or None
            phone = cell(row, idx_phone) or None
            tags_raw = cell(row, idx_tags)
            notes = cell(row, idx_notes) or None

            person, was_created = _find_or_create_person_for_import(full_name, email, phone)
            if was_created:
                created += 1
            else:
                updated += 1
            if notes:
                person.notes = notes
            for tag in [t.strip().lower() for t in tags_raw.replace(';', ',').split(',') if t.strip()]:
                if tag in ('visitor', 'donator', 'member'):
                    person.add_tag(tag)

        db.session.commit()
        log_event('admin_action', f'Spreadsheet import: {created} created, {updated} updated, {skipped} skipped.',
                  actor=current_user)
        flash(f'Import complete: {created} created, {updated} updated, {skipped} skipped.', 'success')
        return redirect(url_for('admin.admin_people'))

    return render_template('admin/import_people.html', columns=IMPORT_COLUMNS)


@admin.route('/admin/import/template.xlsx')
@super_admin_required
def admin_import_template_xlsx():
    import io
    from flask import Response
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = Workbook()
    ws = wb.active
    ws.title = 'People Import'
    ws.append(IMPORT_COLUMNS)
    ws.append(['Jane Doe', 'jane@example.com', '5125551234', 'visitor, donator', 'Met at Sunday feast'])

    for cell in ws[1]:
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill('solid', fgColor='143642')
        cell.alignment = Alignment(horizontal='center')
    for col_cells in ws.columns:
        max_len = max((len(str(c.value or '')) for c in col_cells), default=10)
        ws.column_dimensions[col_cells[0].column_letter].width = min(max_len + 4, 50)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return Response(
        buf.read(),
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': 'attachment; filename=people-import-template.xlsx'},
    )


# ─── Home Program Outreach (SRS §6) ──────────────────────────

@admin.route('/admin/hpo')
@admin_required
def admin_hpo():
    location_filter = request.args.get('location', '')
    favorable_filter = request.args.get('favorable', '')

    query = Person.query.filter(Person.home_program_outreach_participant == True)
    if location_filter:
        query = query.filter(Person.location == location_filter)
    if favorable_filter == 'yes':
        query = query.filter(Person.favorable_contact == True)
    elif favorable_filter == 'no':
        query = query.filter(Person.favorable_contact == False)

    participants = query.order_by(Person.full_name).all()
    return render_template(
        'admin/hpo.html',
        participants=participants,
        locations=HPO_LOCATIONS,
        location_filter=location_filter,
        favorable_filter=favorable_filter,
    )


@admin.route('/admin/people/<int:person_id>/hpo', methods=['POST'])
@admin_required
def admin_hpo_toggle(person_id):
    person = Person.query.get_or_404(person_id)
    action = request.form.get('action', 'add')
    if action == 'add':
        person.home_program_outreach_participant = True
        location = request.form.get('location', '').strip()
        favorable = request.form.get('favorable_contact', 'no') == 'yes'
        if location:
            person.location = location
        person.favorable_contact = favorable
        db.session.commit()
        # Trigger notification email
        if person.email:
            try:
                tmpl = MessageTemplate.query.filter_by(template_type='hpo_added').first()
                if tmpl:
                    from .email_utils import send_email
                    body = tmpl.render({'first_name': person.full_name.split()[0], 'date': datetime.utcnow().strftime('%Y-%m-%d')})
                    send_email(person.email, tmpl.render_subject({'first_name': person.full_name.split()[0]}), body)
            except Exception:
                pass
        flash(f'{person.full_name} added to Home Program Outreach.', 'success')
    else:
        person.home_program_outreach_participant = False
        db.session.commit()
        flash(f'{person.full_name} removed from Home Program Outreach.', 'info')
    return redirect(url_for('admin.admin_hpo'))


@admin.route('/admin/people/<int:person_id>/edit', methods=['GET', 'POST'])
@admin_required
def admin_person_edit(person_id):
    person = Person.query.get_or_404(person_id)
    if request.method == 'POST':
        person.full_name = request.form.get('full_name', '').strip() or person.full_name
        person.email = request.form.get('email', '').strip() or None
        person.phone_number = request.form.get('phone_number', '').strip() or None
        person.location = request.form.get('location', '').strip() or None
        person.notes = request.form.get('notes', '').strip() or None
        person.home_program_outreach_participant = request.form.get('hpo') == 'yes'
        person.favorable_contact = request.form.get('favorable_contact') == 'yes'
        new_tags = request.form.getlist('tags')
        person.status_tags = ','.join(new_tags)
        db.session.commit()
        flash(f'Updated record for {person.full_name}.', 'success')
        return redirect(url_for('admin.admin_people'))

    guest_records = GuestIntake.query.filter_by(person_id=person.id).order_by(GuestIntake.created_at.desc()).all()
    sponsorship_records = Sponsorship.query.filter_by(person_id=person.id).order_by(Sponsorship.created_at.desc()).all()
    return render_template(
        'admin/person_edit.html', person=person, locations=HPO_LOCATIONS,
        guest_records=guest_records, sponsorship_records=sponsorship_records,
    )


# ─── Inventory Module (SRS §9) ───────────────────────────────

@admin.route('/admin/inventory')
@admin_required
def admin_inventory():
    category_filter = request.args.get('category', '')
    query = InventoryItem.query
    if category_filter:
        query = query.filter(InventoryItem.category == category_filter)
    items = query.order_by(InventoryItem.name).all()
    categories = db.session.query(InventoryItem.category).distinct().filter(
        InventoryItem.category.isnot(None)
    ).order_by(InventoryItem.category).all()
    categories = [c[0] for c in categories]
    return render_template('admin/inventory.html', items=items,
                           categories=categories, category_filter=category_filter)


@admin.route('/admin/inventory/new', methods=['POST'])
@admin_required
def admin_inventory_create():
    name = request.form.get('name', '').strip()
    category = request.form.get('category', '').strip()
    quantity = request.form.get('quantity', '0').strip()
    description = request.form.get('description', '').strip()

    if not name:
        flash('Item name is required.', 'warning')
        return redirect(url_for('admin.admin_inventory'))

    try:
        qty = int(quantity)
    except ValueError:
        qty = 0

    item = InventoryItem(
        name=name,
        category=category or None,
        quantity=qty,
        description=description or None,
        created_by_user_id=current_user.id,
    )
    db.session.add(item)
    db.session.flush()

    log = InventoryAuditLog(
        item_id=item.id,
        quantity_before=0,
        quantity_after=qty,
        changed_by_user_id=current_user.id,
        note='Initial creation',
    )
    db.session.add(log)
    db.session.commit()
    flash(f'Added inventory item: {name}.', 'success')
    return redirect(url_for('admin.admin_inventory'))


@admin.route('/admin/inventory/<int:item_id>/edit', methods=['POST'])
@admin_required
def admin_inventory_edit(item_id):
    item = InventoryItem.query.get_or_404(item_id)
    old_qty = item.quantity
    item.name = request.form.get('name', '').strip() or item.name
    item.category = request.form.get('category', '').strip() or None
    item.description = request.form.get('description', '').strip() or None
    try:
        new_qty = int(request.form.get('quantity', old_qty))
    except ValueError:
        new_qty = old_qty

    if new_qty != old_qty:
        log = InventoryAuditLog(
            item_id=item.id,
            quantity_before=old_qty,
            quantity_after=new_qty,
            changed_by_user_id=current_user.id,
            note=request.form.get('note', '').strip() or None,
        )
        db.session.add(log)
    item.quantity = new_qty
    db.session.commit()
    flash(f'Updated {item.name}.', 'success')
    return redirect(url_for('admin.admin_inventory'))


@admin.route('/admin/inventory/<int:item_id>/delete', methods=['POST'])
@admin_required
def admin_inventory_delete(item_id):
    item = InventoryItem.query.get_or_404(item_id)
    name = item.name
    db.session.delete(item)
    db.session.commit()
    flash(f'Deleted inventory item: {name}.', 'success')
    return redirect(url_for('admin.admin_inventory'))


# ─── Temple Fact Sheet (SRS §7) — Super Admin only ───────────

@admin.route('/admin/fact-sheet', methods=['GET', 'POST'])
@super_admin_required
def admin_fact_sheet():
    current_sheet = FactSheet.query.order_by(FactSheet.uploaded_at.desc()).first()

    if request.method == 'POST':
        file = request.files.get('fact_sheet_image')
        if not file or not file.filename:
            flash('Please select an image file.', 'warning')
            return render_template('admin/fact_sheet_upload.html', sheet=current_sheet)

        allowed_mimes = {'image/png': 'image/png', 'image/jpeg': 'image/jpeg', 'image/webp': 'image/webp'}
        mime = file.content_type or ''
        ext = file.filename.rsplit('.', 1)[-1].lower()
        if ext in ('jpg', 'jpeg'):
            mime = 'image/jpeg'
        elif ext == 'png':
            mime = 'image/png'
        elif ext == 'webp':
            mime = 'image/webp'

        if mime not in allowed_mimes:
            flash('Only PNG, JPG, and WEBP images are allowed.', 'warning')
            return render_template('admin/fact_sheet_upload.html', sheet=current_sheet)

        img_bytes = file.read()
        img_b64 = base64.b64encode(img_bytes).decode('utf-8')

        if current_sheet:
            # Replace in-place (only one image stored at a time)
            current_sheet.image_data = img_b64
            current_sheet.image_mime = mime
            current_sheet.image_filename = file.filename
            current_sheet.uploaded_at = datetime.utcnow()
            current_sheet.uploaded_by_user_id = current_user.id
        else:
            new_sheet = FactSheet(
                image_data=img_b64,
                image_mime=mime,
                image_filename=file.filename,
                uploaded_by_user_id=current_user.id,
            )
            db.session.add(new_sheet)

        db.session.commit()
        log_event('admin_action', 'Updated Temple Fact Sheet image.', actor=current_user)
        flash('Temple Fact Sheet updated successfully.', 'success')
        return redirect(url_for('admin.admin_fact_sheet'))

    return render_template('admin/fact_sheet_upload.html', sheet=current_sheet)


# ─── Message Templates (SRS §4.4) ────────────────────────────

@admin.route('/admin/message-templates')
@admin_required
def admin_message_templates():
    templates = MessageTemplate.query.order_by(MessageTemplate.name).all()
    return render_template('admin/message_templates.html', templates=templates)


@admin.route('/admin/message-templates/<int:tmpl_id>/edit', methods=['GET', 'POST'])
@admin_required
def admin_message_template_edit(tmpl_id):
    tmpl = MessageTemplate.query.get_or_404(tmpl_id)
    preview = None

    if request.method == 'POST':
        action = request.form.get('action', 'save')
        if action == 'preview':
            preview_body = request.form.get('body', '')
            preview_subject = request.form.get('subject', '')
            sample = {
                'first_name': 'Arjuna',
                'date': datetime.utcnow().strftime('%Y-%m-%d'),
                'donation_amount': '$151',
                'sponsorship_categories': 'Sunday Feast — Prasadam',
                'form_url': 'https://temple-crm-web.onrender.com/donate',
                'member_list': '1. Arjuna Sharma\n2. Draupadi Patel',
            }
            preview = {
                'subject': preview_subject.replace('{', '{{').replace('}}', '}'),
                'body': preview_body,
            }
            for key, value in sample.items():
                preview['subject'] = preview['subject'].replace('{' + key + '}', value)
                preview['body'] = preview['body'].replace('{' + key + '}', value)
        else:
            tmpl.subject = request.form.get('subject', '').strip()
            tmpl.body = request.form.get('body', '').strip()
            tmpl.updated_at = datetime.utcnow()
            tmpl.updated_by_user_id = current_user.id
            db.session.commit()
            flash(f'Template "{tmpl.name}" saved.', 'success')
            return redirect(url_for('admin.admin_message_templates'))

    return render_template('admin/message_template_edit.html', tmpl=tmpl, preview=preview)


# ─── Payment Settings (SRS §8.3 / §8.6) — Super Admin ───────

@admin.route('/admin/payment-settings', methods=['GET', 'POST'])
@super_admin_required
def admin_payment_settings():
    settings = PaymentSettings.query.first()
    if not settings:
        settings = PaymentSettings()
        db.session.add(settings)
        db.session.commit()

    if request.method == 'POST':
        settings.zelle_details = request.form.get('zelle_details', '').strip() or None
        settings.gpay_details = request.form.get('gpay_details', '').strip() or None
        settings.paypal_details = request.form.get('paypal_details', '').strip() or None
        settings.other_details = request.form.get('other_details', '').strip() or None
        settings.updated_at = datetime.utcnow()
        settings.updated_by_user_id = current_user.id
        db.session.commit()
        log_event('admin_action', 'Updated payment settings.', actor=current_user)
        flash('Payment settings saved.', 'success')
        return redirect(url_for('admin.admin_payment_settings'))

    return render_template('admin/payment_settings.html', settings=settings)


# ─── Security Audit Log — Super Admin only ───────────────────

SECURITY_EVENT_TYPES = [
    'login_success', 'login_failed', 'account_locked', 'logout',
    'record_deleted', 'role_changed', 'admin_action',
]


@admin.route('/admin/security-log')
@super_admin_required
def admin_security_log():
    event_filter = request.args.get('event_type', '')
    query = SecurityAuditLog.query
    if event_filter:
        query = query.filter_by(event_type=event_filter)
    events = query.order_by(SecurityAuditLog.created_at.desc()).limit(300).all()
    return render_template(
        'admin/security_log.html', events=events,
        event_types=SECURITY_EVENT_TYPES, event_filter=event_filter,
    )


# ─── Thursday Broadcast (SRS §4.3) ───────────────────────────

@admin.route('/admin/broadcast', methods=['GET', 'POST'])
@admin_required
def admin_broadcast():
    tmpl = MessageTemplate.query.filter_by(template_type='thursday_broadcast').first()
    sent_count = 0
    if request.method == 'POST':
        mode = request.form.get('mode', 'group')
        if mode == 'individual':
            # Send to every Person with an email address
            people = Person.query.filter(Person.email.isnot(None)).all()
            for person in people:
                try:
                    from .email_utils import send_email
                    ctx = {
                        'first_name': person.full_name.split()[0],
                        'date': datetime.utcnow().strftime('%Y-%m-%d'),
                    }
                    body = tmpl.render(ctx) if tmpl else 'Temple program reminder this Sunday!'
                    subject = tmpl.render_subject(ctx) if tmpl else 'Sunday Temple Program'
                    send_email(person.email, subject, body.replace('\n', '<br>'))
                    sent_count += 1
                except Exception:
                    pass
            flash(f'Broadcast sent to {sent_count} individuals.', 'success')
        else:
            # Group mode — placeholder (WhatsApp API not configured yet)
            flash('Group broadcast queued. WhatsApp delivery will be enabled once API provider is confirmed.', 'info')
    return render_template('admin/broadcast.html', tmpl=tmpl)
