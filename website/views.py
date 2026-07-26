import io
import base64
from datetime import datetime, timedelta

from flask import (Blueprint, Response, current_app, flash, jsonify, redirect,
                   render_template, request, send_from_directory, url_for)
from flask_login import current_user, login_required
from sqlalchemy import desc

from . import db
from .audit import log_event
from .utils import format_phone
from .models import GuestIntake, Sponsorship, Person, FactSheet, PaymentSettings, PendingNewcomer

views = Blueprint('views', __name__)

ROLE_LABELS = {
    'user': 'User',
    'member': 'Member',
    'sales': 'Sales',
    'sunday_school_teacher': 'Sunday School Teacher',
    'congregational_development': 'Congregational Development',
    'admins': 'Admin',
    'super_admin': 'Super Admin',
}

REFERRAL_OPTIONS = [
    'Drive by', 'Social Media', 'Book Rack', 'Friends and Family', 'Google', 'Other',
]

AREA_OPTIONS = [
    'Round Rock', 'Georgetown', 'Leander', 'Cedar Park', 'Liberty Hill',
    'Pflugerville', 'North Austin', 'Downtown', 'South Austin', 'Other',
]

# SRS §3.2 — sponsorship categories
SPONSORSHIP_CATEGORIES = [
    'Sunday Feast — Prasadam',
    'Sunday Feast — Garland',
    'Sunday Feast — Flower',
    'General Donation (one-time)',
    'Special Program / Festival',
]

# SRS §3.3 — payment options (single-select)
PAYMENT_OPTIONS = ['Zelle', 'Google Pay (GPay)', 'PayPal', 'Other', 'Not Yet Paid']

WHATSAPP_OPTIONS = ['Yes', 'No', 'Already on it']
DUPLICATE_WINDOW = timedelta(minutes=10)


def _parse_date(value):
    if not value:
        return None
    try:
        return datetime.strptime(value, '%Y-%m-%d').date()
    except ValueError:
        return None


def _admin_only():
    if not getattr(current_user, 'is_admin', False):
        flash('Reports are only available to admins.', 'warning')
        return False
    return True


def _recent_cutoff():
    return datetime.utcnow() - DUPLICATE_WINDOW


def _find_or_create_person(full_name, email, phone_number, created_by_user_id=None):
    person = None
    if email:
        person = Person.query.filter_by(email=email).first()
    if person is None and phone_number:
        person = Person.query.filter(
            Person.phone_number == phone_number,
            Person.email.is_(None)
        ).first()
    if person is None:
        person = Person(
            full_name=full_name,
            email=email or None,
            phone_number=phone_number or None,
            created_by_user_id=created_by_user_id,
        )
        db.session.add(person)
        db.session.flush()
    return person


def _create_guest_intake_record(full_name, phone_number, email, notes, first_time_at_temple,
                                 whatsapp_status, referral_sources, referral_other,
                                 residence_areas, residence_other, interested_in_volunteering,
                                 janmashtami_setup_help, created_by_user_id):
    """Creates a GuestIntake row and merges it into the unified Person record (visitor tag).
    Shared by the internal staff form and the public-form approval flow."""
    record = GuestIntake(
        full_name=full_name,
        phone_number=phone_number or None,
        email=email or None,
        notes=notes or None,
        first_time_at_temple=first_time_at_temple,
        whatsapp_status=whatsapp_status or None,
        referral_sources=', '.join(referral_sources) if referral_sources else None,
        referral_other=referral_other or None,
        residence_areas=', '.join(residence_areas) if residence_areas else None,
        residence_other=residence_other or None,
        interested_in_volunteering=interested_in_volunteering,
        janmashtami_setup_help=janmashtami_setup_help,
        created_by_user_id=created_by_user_id,
    )
    db.session.add(record)
    db.session.flush()

    person = _find_or_create_person(full_name, email or None, phone_number or None, created_by_user_id)
    person.add_tag('visitor')
    record.person_id = person.id
    db.session.commit()
    return record


def _duplicate_guest_record(full_name, phone_number, email, first_time_value, whatsapp_status,
                             referral_sources, referral_other, residence_areas, residence_other, notes):
    return GuestIntake.query.filter(
        GuestIntake.created_by_user_id == current_user.id,
        GuestIntake.created_at >= _recent_cutoff(),
        GuestIntake.full_name == full_name,
        GuestIntake.phone_number == (phone_number or None),
        GuestIntake.email == (email or None),
        GuestIntake.first_time_at_temple == (first_time_value == 'yes'),
        GuestIntake.whatsapp_status == whatsapp_status,
        GuestIntake.referral_sources == (', '.join(referral_sources) if referral_sources else None),
        GuestIntake.referral_other == (referral_other or None),
        GuestIntake.residence_areas == (', '.join(residence_areas) if residence_areas else None),
        GuestIntake.residence_other == (residence_other or None),
        GuestIntake.notes == (notes or None),
    ).first()


# ─────────────────────────────────────────────────────────────
# Dashboard
# ─────────────────────────────────────────────────────────────

@views.route('/')
@login_required
def home():
    now = datetime.utcnow()
    week_start = now - timedelta(days=7)
    upcoming_cutoff = now.date() + timedelta(days=30)

    my_guest_count = GuestIntake.query.filter_by(created_by_user_id=current_user.id).count()
    my_sponsorship_count = Sponsorship.query.filter_by(created_by_user_id=current_user.id).count()
    newcomers_this_week = GuestIntake.query.filter(GuestIntake.created_at >= week_start).count()
    upcoming_sponsorships = Sponsorship.query.filter(
        Sponsorship.sponsorship_date.isnot(None),
        Sponsorship.sponsorship_date >= now.date(),
        Sponsorship.sponsorship_date <= upcoming_cutoff,
    ).order_by(Sponsorship.sponsorship_date.asc()).limit(5).all()

    return render_template(
        'Home.HTML',
        user=current_user,
        role_labels=ROLE_LABELS,
        metrics={
            'my_guest_count': my_guest_count,
            'my_sponsorship_count': my_sponsorship_count,
            'newcomers_this_week': newcomers_this_week,
        },
        upcoming_sponsorships=upcoming_sponsorships,
    )


# ─────────────────────────────────────────────────────────────
# Newcomer Intake
# ─────────────────────────────────────────────────────────────

@views.route('/newcomers/new', methods=['GET', 'POST'])
@login_required
def guest_intake():
    if request.method == 'POST':
        form = request.form
        full_name = form.get('full_name', '').strip()
        phone_number = format_phone(form.get('phone_number', '').strip())
        email = form.get('email', '').strip()
        notes = form.get('notes', '').strip()
        first_time_value = form.get('first_time_at_temple', 'no')
        whatsapp_status = form.get('whatsapp_status', '').strip()
        referral_sources = form.getlist('referral_sources')
        referral_other = form.get('referral_other', '').strip()
        residence_areas = form.getlist('residence_areas')
        residence_other = form.get('residence_other', '').strip()
        interested_in_volunteering = form.get('interested_in_volunteering') == 'yes'
        janmashtami_setup_help = interested_in_volunteering and form.get('janmashtami_setup_help') == 'yes'

        errors = []
        if not full_name:
            errors.append('Full name is required.')
        if not phone_number:
            errors.append('Phone number is required.')
        if not whatsapp_status:
            errors.append('Please select WhatsApp interest.')

        if errors:
            return render_template(
                'guest-intake.html', errors=errors, form=form,
                referral_options=REFERRAL_OPTIONS, area_options=AREA_OPTIONS,
                whatsapp_options=WHATSAPP_OPTIONS,
            )

        duplicate = _duplicate_guest_record(
            full_name, phone_number, email, first_time_value, whatsapp_status,
            referral_sources, referral_other, residence_areas, residence_other, notes,
        )
        if duplicate:
            flash('This newcomer was already saved. No duplicate record was created.', 'info')
            if form.get('submit_action') == 'save_add_another':
                return redirect(url_for('views.guest_intake'))
            return redirect(url_for('views.my_data'))

        _create_guest_intake_record(
            full_name, phone_number, email, notes, (first_time_value == 'yes'),
            whatsapp_status, referral_sources, referral_other, residence_areas, residence_other,
            interested_in_volunteering, janmashtami_setup_help, current_user.id,
        )

        flash('Guest intake saved successfully.', 'success')
        if form.get('submit_action') == 'save_add_another':
            return redirect(url_for('views.guest_intake'))
        return redirect(url_for('views.my_data'))

    return render_template(
        'guest-intake.html',
        referral_options=REFERRAL_OPTIONS,
        area_options=AREA_OPTIONS,
        whatsapp_options=WHATSAPP_OPTIONS,
    )


# ─────────────────────────────────────────────────────────────
# Public Newcomer Signup Form — pending admin approval
# ─────────────────────────────────────────────────────────────

@views.route('/newcomer-signup', methods=['GET', 'POST'])
def newcomer_signup():
    if request.method == 'POST':
        form = request.form
        full_name = form.get('full_name', '').strip()
        phone_number = format_phone(form.get('phone_number', '').strip())
        email = form.get('email', '').strip()
        notes = form.get('notes', '').strip()
        first_time_value = form.get('first_time_at_temple', 'no')
        whatsapp_status = form.get('whatsapp_status', '').strip()
        referral_sources = form.getlist('referral_sources')
        referral_other = form.get('referral_other', '').strip()
        residence_areas = form.getlist('residence_areas')
        residence_other = form.get('residence_other', '').strip()
        interested_in_volunteering = form.get('interested_in_volunteering') == 'yes'
        janmashtami_setup_help = interested_in_volunteering and form.get('janmashtami_setup_help') == 'yes'

        errors = []
        if not full_name:
            errors.append('Full name is required.')
        if not phone_number:
            errors.append('Phone number is required.')
        if not whatsapp_status:
            errors.append('Please select your WhatsApp interest.')

        if errors:
            return render_template(
                'newcomer-signup.html', errors=errors, form=form,
                referral_options=REFERRAL_OPTIONS, area_options=AREA_OPTIONS,
                whatsapp_options=WHATSAPP_OPTIONS,
            )

        pending = PendingNewcomer(
            full_name=full_name,
            phone_number=phone_number or None,
            email=email or None,
            notes=notes or None,
            first_time_at_temple=(first_time_value == 'yes'),
            whatsapp_status=whatsapp_status,
            referral_sources=', '.join(referral_sources) if referral_sources else None,
            referral_other=referral_other or None,
            residence_areas=', '.join(residence_areas) if residence_areas else None,
            residence_other=residence_other or None,
            interested_in_volunteering=interested_in_volunteering,
            janmashtami_setup_help=janmashtami_setup_help,
        )
        db.session.add(pending)
        db.session.commit()

        return render_template('newcomer-signup.html', success=True)

    return render_template(
        'newcomer-signup.html',
        referral_options=REFERRAL_OPTIONS,
        area_options=AREA_OPTIONS,
        whatsapp_options=WHATSAPP_OPTIONS,
    )


@views.route('/newcomer-signup/qr.png')
def newcomer_signup_qr():
    try:
        import qrcode
        form_url = request.host_url.rstrip('/') + url_for('views.newcomer_signup')
        qr = qrcode.QRCode(version=1, box_size=10, border=4)
        qr.add_data(form_url)
        qr.make(fit=True)
        img = qr.make_image(fill_color='#143642', back_color='white')
        buf = io.BytesIO()
        img.save(buf, format='PNG')
        buf.seek(0)
        return Response(buf.read(), mimetype='image/png')
    except ImportError:
        return Response('qrcode library not installed', status=500)


# ─────────────────────────────────────────────────────────────
# Sponsorship (staff form)
# ─────────────────────────────────────────────────────────────

@views.route('/sponsorships/new', methods=['GET', 'POST'])
@login_required
def sponsorship():
    if request.method == 'POST':
        form = request.form
        sponsor_legal_name = form.get('sponsor_legal_name', '').strip()
        sponsor_spiritual_name = form.get('sponsor_spiritual_name', '').strip()
        sponsor_phone_number = form.get('sponsor_phone_number', '').strip()
        sponsor_email = form.get('sponsor_email', '').strip()
        sponsorship_cats = form.getlist('sponsorship_categories')
        occasion = form.get('occasion', '').strip()
        sponsorship_date_raw = form.get('sponsorship_date', '').strip()
        sponsorship_date = _parse_date(sponsorship_date_raw)
        donation_amount = form.get('donation_amount', '').strip()
        notes = form.get('notes', '').strip()
        payment_method = form.get('payment_method', '').strip()
        payment_method_other = form.get('payment_method_other', '').strip()

        errors = []
        if not sponsor_legal_name:
            errors.append('Sponsor legal name is required.')
        if not sponsor_phone_number:
            errors.append('Sponsor phone number is required.')
        elif not sponsor_phone_number.isdigit():
            errors.append('Sponsor phone number must contain numbers only.')
        if not sponsorship_date_raw:
            errors.append('Date of sponsorship is required.')
        elif not sponsorship_date:
            errors.append('Date of sponsorship must be a valid date.')
        if not sponsorship_cats:
            errors.append('Select at least one sponsorship category.')
        if not payment_method:
            errors.append('Select a payment method.')

        if errors:
            return render_template(
                'sponsorship.html', errors=errors, form=form,
                sponsorship_categories=SPONSORSHIP_CATEGORIES,
                payment_options=PAYMENT_OPTIONS,
            )

        sponsor_phone_number = format_phone(sponsor_phone_number)
        not_yet_paid = (payment_method == 'Not Yet Paid')

        record = Sponsorship(
            sponsor_legal_name=sponsor_legal_name,
            sponsor_spiritual_name=sponsor_spiritual_name or None,
            sponsor_phone_number=sponsor_phone_number or None,
            sponsor_email=sponsor_email or None,
            sponsorship_categories=', '.join(sponsorship_cats) if sponsorship_cats else None,
            occasion=occasion or None,
            sponsorship_date=sponsorship_date,
            donation_amount=donation_amount or None,
            notes=notes or None,
            payment_methods=payment_method or None,
            payment_method_other=payment_method_other or None,
            not_yet_paid=not_yet_paid,
            source='internal',
            created_by_user_id=current_user.id,
        )
        db.session.add(record)
        db.session.flush()

        # Create/update unified Person record with donator tag
        person = _find_or_create_person(
            sponsor_legal_name, sponsor_email or None, sponsor_phone_number or None, current_user.id
        )
        person.add_tag('donator')
        record.person_id = person.id
        db.session.commit()

        # Send Not Yet Paid email if applicable
        if not_yet_paid and sponsor_email:
            _send_not_yet_paid(sponsor_email, sponsor_legal_name)

        flash('Sponsorship saved successfully.', 'success')
        return redirect(url_for('views.my_data'))

    return render_template(
        'sponsorship.html',
        sponsorship_categories=SPONSORSHIP_CATEGORIES,
        payment_options=PAYMENT_OPTIONS,
    )


@views.route('/api/people/search')
@login_required
def api_people_search():
    q = request.args.get('q', '').strip()
    if len(q) < 3:
        return jsonify([])
    like = f'%{q}%'
    matches = Person.query.filter(
        db.or_(Person.full_name.ilike(like), Person.email.ilike(like), Person.phone_number.ilike(like))
    ).order_by(Person.full_name).limit(8).all()
    return jsonify([
        {
            'id': p.id,
            'full_name': p.full_name,
            'email': p.email or '',
            'phone_number': p.phone_number or '',
            'tags': p.get_tags(),
        }
        for p in matches
    ])


def _send_not_yet_paid(email, name):
    try:
        from .email_utils import send_not_yet_paid_email
        settings = PaymentSettings.query.first()
        form_url = request.host_url.rstrip('/') + url_for('views.donate')
        send_not_yet_paid_email(email, name, settings, form_url)
    except Exception:
        pass


# ─────────────────────────────────────────────────────────────
# My Data
# ─────────────────────────────────────────────────────────────

@views.route('/my-data')
@login_required
def my_data():
    guest_records = GuestIntake.query.filter_by(
        created_by_user_id=current_user.id
    ).order_by(desc(GuestIntake.created_at)).all()
    sponsorship_records = Sponsorship.query.filter_by(
        created_by_user_id=current_user.id
    ).order_by(desc(Sponsorship.created_at)).all()
    return render_template('my-data.html', guest_records=guest_records, sponsorship_records=sponsorship_records)


# ─────────────────────────────────────────────────────────────
# Admin delete actions
# ─────────────────────────────────────────────────────────────

@views.route('/records/newcomers/<int:record_id>/delete', methods=['POST'])
@login_required
def delete_guest_record(record_id):
    if not _admin_only():
        return redirect(url_for('views.home'))
    record = GuestIntake.query.get_or_404(record_id)
    guest_name = record.full_name
    record_id = record.id
    db.session.delete(record)
    db.session.commit()
    log_event('record_deleted', f'Deleted newcomer record for {guest_name}.', actor=current_user,
              target_type='GuestIntake', target_id=record_id)
    flash(f'Deleted newcomer record for {guest_name}.', 'success')
    return redirect(url_for('views.reports'))


@views.route('/records/sponsorships/<int:record_id>/delete', methods=['POST'])
@login_required
def delete_sponsorship_record(record_id):
    if not _admin_only():
        return redirect(url_for('views.home'))
    record = Sponsorship.query.get_or_404(record_id)
    sponsor_name = record.sponsor_legal_name
    record_id = record.id
    db.session.delete(record)
    db.session.commit()
    log_event('record_deleted', f'Deleted sponsorship record for {sponsor_name}.', actor=current_user,
              target_type='Sponsorship', target_id=record_id)
    flash(f'Deleted sponsorship record for {sponsor_name}.', 'success')
    return redirect(url_for('views.reports'))


# ─────────────────────────────────────────────────────────────
# Reports & Exports
# ─────────────────────────────────────────────────────────────

@views.route('/reports')
@login_required
def reports():
    if not _admin_only():
        return redirect(url_for('views.home'))

    tag_filter = request.args.get('tag', '')
    payment_filter = request.args.get('payment', '')

    guest_records = GuestIntake.query.order_by(desc(GuestIntake.created_at)).all()

    sp_query = Sponsorship.query
    if payment_filter:
        if payment_filter == 'not_yet_paid':
            sp_query = sp_query.filter(Sponsorship.not_yet_paid == True)
        else:
            sp_query = sp_query.filter(Sponsorship.payment_methods == payment_filter)
    sponsorship_records = sp_query.order_by(
        Sponsorship.sponsorship_date.asc(), desc(Sponsorship.created_at)
    ).all()

    whatsapp_ready = GuestIntake.query.filter(
        GuestIntake.whatsapp_status.in_(['Yes', 'Already on it'])
    ).all()

    return render_template(
        'reports.html',
        guest_records=guest_records,
        sponsorship_records=sponsorship_records,
        whatsapp_ready=whatsapp_ready,
        payment_filter=payment_filter,
        payment_options=PAYMENT_OPTIONS,
    )


@views.route('/reports/export/<report_name>.xlsx')
@login_required
def export_report_xlsx(report_name):
    if not _admin_only():
        return redirect(url_for('views.home'))

    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        flash('openpyxl is not installed. Run: pip install openpyxl', 'danger')
        return redirect(url_for('views.reports'))

    wb = openpyxl.Workbook()
    ws = wb.active

    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill('solid', fgColor='143642')

    if report_name == 'newcomers':
        ws.title = 'Newcomers'
        headers = [
            'Full Name', 'Phone', 'Email', 'First Time', 'WhatsApp',
            'Referral Sources', 'Referral Other', 'Residence Areas', 'Residence Other',
            'Interested in Volunteering', 'Janmashtami Setup Help',
            'Notes', 'Recorded By', 'Created At',
        ]
        ws.append(headers)
        for record in GuestIntake.query.order_by(desc(GuestIntake.created_at)).all():
            ws.append([
                record.full_name,
                record.phone_number or '',
                record.email or '',
                'Yes' if record.first_time_at_temple else 'No',
                record.whatsapp_status or '',
                record.referral_sources or '',
                record.referral_other or '',
                record.residence_areas or '',
                record.residence_other or '',
                'Yes' if record.interested_in_volunteering else 'No',
                'Yes' if record.janmashtami_setup_help else 'No',
                record.notes or '',
                record.created_by_user.display_name if record.created_by_user else '',
                record.created_at.strftime('%Y-%m-%d %H:%M'),
            ])

    elif report_name == 'sponsorships':
        ws.title = 'Sponsorships'
        headers = [
            'Sponsor Legal Name', 'Spiritual Name', 'Phone', 'Email',
            'Sponsorship Categories', 'Occasion', 'Date',
            'Amount', 'Payment Method', 'Not Yet Paid',
            'Notes', 'Recorded By', 'Source', 'Created At',
        ]
        ws.append(headers)
        for record in Sponsorship.query.order_by(
            Sponsorship.sponsorship_date.asc(), desc(Sponsorship.created_at)
        ).all():
            ws.append([
                record.sponsor_legal_name,
                record.sponsor_spiritual_name or '',
                record.sponsor_phone_number or '',
                record.sponsor_email or '',
                record.sponsorship_categories or record.sponsoring_for or '',
                record.occasion or '',
                record.sponsorship_date.isoformat() if record.sponsorship_date else '',
                record.donation_amount or record.amount_options or record.amount_other or '',
                record.payment_methods or '',
                'Yes' if record.not_yet_paid else 'No',
                record.notes or '',
                record.created_by_user.display_name if record.created_by_user else 'Public Form',
                record.source or 'internal',
                record.created_at.strftime('%Y-%m-%d %H:%M'),
            ])

    elif report_name == 'people':
        ws.title = 'People'
        headers = ['Full Name', 'Email', 'Phone', 'Status Tags', 'HPO Participant',
                   'Favorable Contact', 'Location', 'Notes', 'Created At']
        ws.append(headers)
        for p in Person.query.order_by(Person.full_name).all():
            ws.append([
                p.full_name, p.email or '', p.phone_number or '',
                p.status_tags or '',
                'Yes' if p.home_program_outreach_participant else 'No',
                'Yes' if p.favorable_contact else 'No',
                p.location or '', p.notes or '',
                p.created_at.strftime('%Y-%m-%d %H:%M'),
            ])
    else:
        flash('Unknown report requested.', 'warning')
        return redirect(url_for('views.reports'))

    # Style header row
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')

    # Auto-size columns
    for col in ws.columns:
        max_len = max((len(str(cell.value or '')) for cell in col), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 50)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    return Response(
        buf.read(),
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': f'attachment; filename={report_name}-report.xlsx'},
    )


# Keep legacy CSV route for backward compat
@views.route('/reports/export/<report_name>.csv')
@login_required
def export_report(report_name):
    return redirect(url_for('views.export_report_xlsx', report_name=report_name))


# ─────────────────────────────────────────────────────────────
# Public Donation Report Form (SRS §8)
# ─────────────────────────────────────────────────────────────

@views.route('/donate', methods=['GET', 'POST'])
def donate():
    payment_settings = PaymentSettings.query.first()

    if request.method == 'POST':
        form = request.form
        full_name = form.get('full_name', '').strip()
        email = form.get('email', '').strip()
        phone_number = format_phone(form.get('phone_number', '').strip())
        sponsorship_cats = form.getlist('sponsorship_categories')
        payment_method = form.get('payment_method', '').strip()
        donation_amount = form.get('donation_amount', '').strip()
        notes = form.get('notes', '').strip()

        errors = []
        if not full_name:
            errors.append('Full name is required.')
        if not email:
            errors.append('Email address is required.')
        if not sponsorship_cats:
            errors.append('Please select at least one sponsorship category.')
        if not payment_method:
            errors.append('Please select a payment method.')
        if not donation_amount:
            errors.append('Donation amount is required.')

        if errors:
            return render_template(
                'donate.html', errors=errors, form=form,
                sponsorship_categories=SPONSORSHIP_CATEGORIES,
                payment_options=PAYMENT_OPTIONS,
                payment_settings=payment_settings,
            )

        not_yet_paid = (payment_method == 'Not Yet Paid')

        record = Sponsorship(
            sponsor_legal_name=full_name,
            sponsor_email=email or None,
            sponsor_phone_number=phone_number or None,
            sponsorship_categories=', '.join(sponsorship_cats),
            donation_amount=donation_amount,
            payment_methods=payment_method,
            not_yet_paid=not_yet_paid,
            notes=notes or None,
            source='public',
            created_by_user_id=None,
        )
        db.session.add(record)
        db.session.flush()

        # Create/update unified Person record
        person = _find_or_create_person(full_name, email or None, phone_number or None)
        person.add_tag('donator')
        record.person_id = person.id
        db.session.commit()

        # Send emails
        if not_yet_paid and email:
            try:
                from .email_utils import send_not_yet_paid_email
                form_url = request.host_url.rstrip('/') + url_for('views.donate')
                send_not_yet_paid_email(email, full_name, payment_settings, form_url)
            except Exception:
                pass
        elif email:
            try:
                from .email_utils import send_donation_confirmation_email
                send_donation_confirmation_email(
                    email, full_name,
                    ', '.join(sponsorship_cats),
                    donation_amount,
                )
            except Exception:
                pass

        return render_template('donate.html', success=True)

    return render_template(
        'donate.html',
        sponsorship_categories=SPONSORSHIP_CATEGORIES,
        payment_options=PAYMENT_OPTIONS,
        payment_settings=payment_settings,
    )


# ─────────────────────────────────────────────────────────────
# Temple Fact Sheet — public view (SRS §7)
# ─────────────────────────────────────────────────────────────

@views.route('/fact-sheet')
def fact_sheet():
    sheet = FactSheet.query.order_by(desc(FactSheet.uploaded_at)).first()
    return render_template('fact-sheet.html', sheet=sheet)


@views.route('/fact-sheet/image')
def fact_sheet_image():
    sheet = FactSheet.query.order_by(desc(FactSheet.uploaded_at)).first()
    if not sheet:
        return Response('No image uploaded', status=404)
    img_bytes = base64.b64decode(sheet.image_data)
    return Response(img_bytes, mimetype=sheet.image_mime)


# ─────────────────────────────────────────────────────────────
# QR Code for donation form (SRS §5.4)
# ─────────────────────────────────────────────────────────────

@views.route('/donate/qr.png')
def donate_qr():
    try:
        import qrcode
        donate_url = request.host_url.rstrip('/') + url_for('views.donate')
        qr = qrcode.QRCode(version=1, box_size=10, border=4)
        qr.add_data(donate_url)
        qr.make(fit=True)
        img = qr.make_image(fill_color='#143642', back_color='white')
        buf = io.BytesIO()
        img.save(buf, format='PNG')
        buf.seek(0)
        return Response(buf.read(), mimetype='image/png')
    except ImportError:
        return Response('qrcode library not installed', status=500)


# ─────────────────────────────────────────────────────────────
# PWA helpers
# ─────────────────────────────────────────────────────────────

@views.route('/manifest.json')
def manifest():
    return send_from_directory(current_app.static_folder, 'manifest.json')


@views.route('/sw.js')
def service_worker():
    response = send_from_directory(current_app.static_folder, 'sw.js')
    response.headers['Cache-Control'] = 'no-cache'
    return response


@views.route('/offline')
def offline():
    return render_template('offline.html')
